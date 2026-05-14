"""
Extração de Notas Fiscais de Serviços (NFSe) de PDF judicial.
Processa páginas 816-5840 e gera Excel com todos os dados.
"""
import fitz  # PyMuPDF
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import time
import sys
from pathlib import Path

PDF_PATH = Path(__file__).parent / "0826637-88.2026.8.10.0001.pdf"
OUTPUT_PATH = Path(__file__).parent / "notas_fiscais.xlsx"

START_PAGE = 815   # 0-indexed (página 816)
END_PAGE   = 5839  # 0-indexed inclusive (página 5840)

# Regex para valor monetário brasileiro: 1.234,56 ou 234,56
RE_VALOR = re.compile(r'[\d.]+,\d{2}')
# Regex para data DD/MM/YYYY
RE_DATA  = re.compile(r'\d{2}/\d{2}/\d{4}')
# Regex para CNPJ
RE_CNPJ  = re.compile(r'\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}')
# Regex para item de NFSe no bloco PyMuPDF:
# bloco contém: quantidade\nvalor_unit\nvalor_total\nTRIBUT...\ndescricao
RE_ITEM  = re.compile(
    r'^(\d+)\n([\d.]+,\d{2})\n([\d.]+,\d{2})\n(TRIBUT[^\n]+)\n(.+?)(?:\n|$)',
    re.DOTALL | re.IGNORECASE
)
# Regex para valor total da nota (R$ como qualquer char após R)
RE_TOTAL = re.compile(r'VALOR TOTAL DA NOTA\s*=\s*R.\s*([\d.]+,\d{2})', re.IGNORECASE)


def parse_nfse(blocks: list) -> dict:
    """Extrai campos de uma NFSe a partir dos blocos PyMuPDF da página."""
    sorted_blocks = sorted(blocks, key=lambda b: (b[1], b[0]))  # ordena por Y, X

    nf_number      = ''
    data_emissao   = ''
    nome_emitente  = ''
    cnpj_emitente  = ''
    nome_tomador   = ''
    cnpj_tomador   = ''
    discriminacao  = ''
    items          = []
    valor_total    = ''
    cancelada      = False

    after_tomador_header   = False
    tomador_name_found     = False
    after_cpf_tomador_lbl  = False

    for block in sorted_blocks:
        x0, y0 = block[0], block[1]
        text = block[4].strip()
        if not text:
            continue
        tu = text.upper()

        # ── Cancelamento ───────────────────────────────────────────────
        if 'CANCELADA' in tu:
            cancelada = True

        # ── Número da NF (bloco à direita, y < 30, só dígitos) ─────────
        if y0 < 30 and x0 > 300 and re.match(r'^\d+$', text):
            nf_number = text.lstrip('0') or '0'

        # ── Data de emissão (bloco à direita, y < 60, começa com data) ─
        if y0 < 60 and x0 > 300:
            m = RE_DATA.match(text)
            if m:
                data_emissao = m.group(0)

        # ── Emitente: bloco com "Nome / Razão Social:" e "CPF / CNPJ:" ─
        # O bloco do prestador contém ambos os labels num único bloco.
        if 'NOME / RAZ' in tu and 'CPF / CNPJ:' in text:
            lines = text.split('\n')
            for i, line in enumerate(lines):
                if 'NOME / RAZ' in line.upper() and i + 1 < len(lines):
                    nome_emitente = lines[i + 1].strip()
                if 'CPF / CNPJ:' in line.upper() and i + 1 < len(lines):
                    cnpj_emitente = lines[i + 1].strip()

        # ── Tomador ────────────────────────────────────────────────────
        if 'TOMADOR DE SERVI' in tu:
            after_tomador_header  = True
            tomador_name_found    = False
            after_cpf_tomador_lbl = False

        # Nome do tomador: primeiro bloco após header com x0 > 100 sem label
        if after_tomador_header and not tomador_name_found:
            skip_labels = ('NOME', 'CPF', 'ENDERE', 'MUNIC', 'DISCRIMINA',
                           'INSCRI', 'TELEFONE', 'EMAIL', 'UF:', 'SAO LUIS')
            if x0 > 100 and not any(kw in tu for kw in skip_labels):
                nome_tomador   = text.split('\n')[0].strip()
                tomador_name_found = True

        # CNPJ do tomador: "CPF/CNPJ:" sem espaços (diferencia do emitente)
        if 'CPF/CNPJ:' in text and 'CPF / CNPJ:' not in text:
            m = RE_CNPJ.search(text)
            if m:
                cnpj_tomador = m.group(0)
                after_cpf_tomador_lbl = False
            else:
                after_cpf_tomador_lbl = True

        elif after_cpf_tomador_lbl:
            m = RE_CNPJ.search(text)
            if m:
                cnpj_tomador = m.group(0)
                after_cpf_tomador_lbl = False

        # ── Discriminação dos serviços ─────────────────────────────────
        if not discriminacao and tu.startswith('DESCRI') and ':' in text:
            m = re.search(r'Descri[çc][ãa]o:(.+)', text, re.IGNORECASE)
            if m:
                discriminacao = m.group(1).strip().split('\n')[0]

        # ── Itens da nota ──────────────────────────────────────────────
        m = RE_ITEM.match(text)
        if m:
            items.append({
                'descricao_item':  m.group(5).split('\n')[0].strip(),
                'quantidade':      m.group(1),
                'valor_unitario':  m.group(2),
                'valor_total_item': m.group(3),
            })

        # ── Valor total da nota ────────────────────────────────────────
        m = RE_TOTAL.search(text)
        if m:
            valor_total = m.group(1)

    if not items:
        items = [{'descricao_item': '', 'quantidade': '', 'valor_unitario': '', 'valor_total_item': ''}]

    return {
        'numero_nf':     nf_number,
        'data_emissao':  data_emissao,
        'nome_emitente': nome_emitente,
        'cnpj_emitente': cnpj_emitente,
        'nome_tomador':  nome_tomador,
        'cnpj_tomador':  cnpj_tomador,
        'discriminacao': discriminacao,
        'items':         items,
        'valor_total':   valor_total,
        'cancelada':     'SIM' if cancelada else 'NÃO',
    }


def build_excel(rows: list) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Notas Fiscais'

    headers = [
        'Página PDF', 'Nº NF', 'Data Emissão',
        'CNPJ Emitente', 'Nome Emitente',
        'CNPJ Tomador', 'Nome Tomador',
        'Discriminação dos Serviços',
        'Descrição do Item', 'Quantidade',
        'Valor Unitário (R$)', 'Valor Total Item (R$)',
        'Valor Total NF (R$)', 'Cancelada',
    ]
    ws.append(headers)

    # Estilo do cabeçalho
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(bold=True, color='FFFFFF', size=10)
    thin     = Side(style='thin', color='AAAAAA')
    border   = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
    ws.row_dimensions[1].height = 30

    cancel_fill   = PatternFill('solid', fgColor='FFCCCC')
    cancel_font   = Font(color='990000', size=9)
    normal_font   = Font(size=9)
    center_align  = Alignment(horizontal='center', vertical='center')
    left_align    = Alignment(horizontal='left', vertical='center', wrap_text=True)

    for row_data in rows:
        ws.append(row_data['cells'])
        row_idx = ws.max_row
        is_cancelled = row_data['cancelada']
        for i, cell in enumerate(ws[row_idx]):
            cell.border = border
            if is_cancelled:
                cell.fill = cancel_fill
                cell.font = cancel_font
            else:
                cell.font = normal_font
            # Colunas numéricas centralizadas
            if i in (0, 1, 9):
                cell.alignment = center_align
            else:
                cell.alignment = left_align

    # Larguras das colunas
    col_widths = [10, 10, 13, 22, 42, 22, 42, 55, 35, 12, 20, 20, 20, 11]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    return wb


def main():
    print(f'Abrindo PDF: {PDF_PATH}')
    doc = fitz.open(str(PDF_PATH))
    total_pages = len(doc)
    print(f'Total de páginas: {total_pages}')
    print(f'Processando páginas {START_PAGE + 1} a {END_PAGE + 1}...\n')

    rows       = []
    nf_count   = 0
    skip_count = 0
    t0         = time.time()

    for pg_num in range(START_PAGE, min(END_PAGE + 1, total_pages)):
        page   = doc[pg_num]
        blocks = page.get_text('blocks')

        # Verificação rápida antes de parsear
        all_text = ' '.join(b[4] for b in blocks)
        if 'NOTA FISCAL DE SERVI' not in all_text.upper():
            skip_count += 1
            continue

        data     = parse_nfse(blocks)
        nf_count += 1

        is_cancelled = data['cancelada'] == 'SIM'

        for item in data['items']:
            rows.append({
                'cancelada': is_cancelled,
                'cells': [
                    pg_num + 1,
                    data['numero_nf'],
                    data['data_emissao'],
                    data['cnpj_emitente'],
                    data['nome_emitente'],
                    data['cnpj_tomador'],
                    data['nome_tomador'],
                    data['discriminacao'],
                    item['descricao_item'],
                    item['quantidade'],
                    item['valor_unitario'],
                    item['valor_total_item'],
                    data['valor_total'],
                    data['cancelada'],
                ],
            })

        if nf_count % 500 == 0:
            elapsed   = time.time() - t0
            pgs_done  = pg_num - START_PAGE + 1
            pgs_total = END_PAGE - START_PAGE + 1
            pct       = pgs_done / pgs_total * 100
            eta       = elapsed / pgs_done * (pgs_total - pgs_done)
            print(f'  {nf_count} notas | pág {pg_num + 1} | {pct:.1f}% | ETA {eta:.0f}s', flush=True)

    doc.close()

    elapsed = time.time() - t0
    print(f'\nExtração concluída: {nf_count} notas, {len(rows)} linhas ({elapsed:.1f}s)')
    print(f'Páginas puladas (sem NF): {skip_count}')
    print('Gerando Excel...')

    wb = build_excel(rows)
    wb.save(str(OUTPUT_PATH))
    print(f'Salvo em: {OUTPUT_PATH}')


if __name__ == '__main__':
    main()
